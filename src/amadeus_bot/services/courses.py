from __future__ import annotations

import csv
import io
import json
import os
import re
import secrets
from dataclasses import dataclass
from datetime import datetime, time, timedelta
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from amadeus_bot.repositories.user_data import UserDataRepository


@dataclass(frozen=True, slots=True)
class CourseRecord:
    course_id: int
    name: str
    weekday: int
    start_section: int
    end_section: int
    weeks: str
    location: str
    teacher: str
    reminder_minutes: int | None


@dataclass(slots=True)
class ImportPreview:
    token: str
    user_id: str
    rows: list[dict[str, Any]]
    source: str


class CourseService:
    COLUMN_ALIASES = {
        "name": {"课程名", "课程名称", "课程", "name"},
        "teacher": {"教师", "老师", "任课教师", "teacher"},
        "location": {"教室", "地点", "上课地点", "location"},
        "weekday": {"星期", "周几", "星期几", "weekday"},
        "start_section": {"开始节次", "起始节次", "开始节", "start_section"},
        "end_section": {"结束节次", "终止节次", "结束节", "end_section"},
        "weeks": {"周次", "起止周", "教学周", "weeks"},
        "campus": {"校区", "campus"},
    }

    def __init__(self, repository: UserDataRepository) -> None:
        self.repository = repository
        self._previews: dict[str, ImportPreview] = {}

    def add(
        self,
        user_id: str,
        name: str,
        weekday: int,
        sections: tuple[int, int],
        weeks: str,
        location: str = "",
        teacher: str = "",
        *,
        source: str = "manual",
        batch_id: str | None = None,
    ) -> int:
        if not name.strip() or not 1 <= weekday <= 7:
            raise ValueError("课程名不能为空，星期必须为 1～7")
        start, end = sections
        if not 1 <= start <= end <= 20:
            raise ValueError("节次必须在 1～20 且结束节不早于开始节")
        normalize_weeks(weeks)
        with self.repository.connection(user_id) as connection:
            cursor = connection.execute(
                """
                INSERT INTO courses(
                    name,teacher,location,weekday,start_section,end_section,weeks,source,import_batch_id
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    name.strip(),
                    teacher.strip(),
                    location.strip(),
                    weekday,
                    start,
                    end,
                    weeks,
                    source,
                    batch_id,
                ),
            )
            return int(cursor.lastrowid)

    def list(self, user_id: str, weekday: int | None = None) -> list[CourseRecord]:
        with self.repository.connection(user_id) as connection:
            if weekday is None:
                rows = connection.execute(
                    "SELECT * FROM courses WHERE deleted_at_utc IS NULL ORDER BY weekday,start_section"
                ).fetchall()
            else:
                rows = connection.execute(
                    """
                    SELECT * FROM courses WHERE deleted_at_utc IS NULL AND weekday=?
                    ORDER BY start_section
                    """,
                    (weekday,),
                ).fetchall()
        return [self._from_row(row) for row in rows]

    def edit(self, user_id: str, course_id: int, changes: dict[str, Any]) -> bool:
        allowed = {"name", "teacher", "location", "weekday", "start_section", "end_section", "weeks"}
        values = {key: value for key, value in changes.items() if key in allowed}
        if not values:
            raise ValueError("没有可修改字段")
        assignments = ",".join(f"{key}=?" for key in values)
        with self.repository.connection(user_id) as connection:
            cursor = connection.execute(
                f"UPDATE courses SET {assignments},updated_at_utc=CURRENT_TIMESTAMP "
                "WHERE course_id=? AND deleted_at_utc IS NULL",
                (*values.values(), int(course_id)),
            )
            return cursor.rowcount > 0

    def delete(self, user_id: str, course_id: int) -> bool:
        with self.repository.connection(user_id) as connection:
            cursor = connection.execute(
                """
                UPDATE courses SET deleted_at_utc=CURRENT_TIMESTAMP
                WHERE course_id=? AND deleted_at_utc IS NULL
                """,
                (int(course_id),),
            )
            return cursor.rowcount > 0

    def set_reminder(self, user_id: str, course_id: int | None, minutes: int | None) -> int:
        if minutes is not None and not 0 <= minutes <= 1440:
            raise ValueError("提前分钟数必须在 0～1440")
        with self.repository.connection(user_id) as connection:
            if course_id is None:
                cursor = connection.execute(
                    "UPDATE courses SET reminder_minutes=? WHERE deleted_at_utc IS NULL", (minutes,)
                )
            else:
                cursor = connection.execute(
                    """
                    UPDATE courses SET reminder_minutes=?
                    WHERE course_id=? AND deleted_at_utc IS NULL
                    """,
                    (minutes, int(course_id)),
                )
            return cursor.rowcount

    def due_reminders(self, *, now: datetime | None = None) -> list[tuple[str, CourseRecord, str]]:
        semester_start_raw = os.getenv("AMADEUS_SEMESTER_START", "").strip()
        if not semester_start_raw or not self.repository.users_root.exists():
            return []
        zone = ZoneInfo("Asia/Shanghai")
        current = (now or datetime.now(zone)).astimezone(zone)
        semester_start = datetime.fromisoformat(semester_start_raw).date()
        week = ((current.date() - semester_start).days // 7) + 1
        if week < 1 or week > 30:
            return []
        section_times = _section_times()
        result: list[tuple[str, CourseRecord, str]] = []
        for user_dir in self.repository.users_root.iterdir():
            if not user_dir.is_dir() or not user_dir.name.isdigit():
                continue
            for course in self.list(user_dir.name, current.isoweekday()):
                if course.reminder_minutes is None or week not in normalize_weeks(course.weeks):
                    continue
                start_text = section_times.get(course.start_section)
                if start_text is None:
                    continue
                start_at = datetime.combine(current.date(), start_text, zone)
                remind_at = start_at - timedelta(minutes=course.reminder_minutes)
                if remind_at <= current < remind_at + timedelta(seconds=45):
                    occurrence = current.date().isoformat()
                    with self.repository.connection(user_dir.name) as connection:
                        existing = connection.execute(
                            """
                            SELECT 1 FROM course_reminder_deliveries
                            WHERE course_id=? AND occurrence_date=?
                            """,
                            (course.course_id, occurrence),
                        ).fetchone()
                    if not existing:
                        result.append((user_dir.name, course, occurrence))
        return result

    def mark_reminder_sent(self, user_id: str, course_id: int, occurrence_date: str) -> bool:
        with self.repository.connection(user_id) as connection:
            cursor = connection.execute(
                """
                INSERT OR IGNORE INTO course_reminder_deliveries(course_id,occurrence_date,sent_at_utc)
                VALUES (?, ?, CURRENT_TIMESTAMP)
                """,
                (course_id, occurrence_date),
            )
            return cursor.rowcount > 0

    def preview_csv(self, user_id: str, content: bytes, filename: str) -> ImportPreview:
        text = content.decode("utf-8-sig")
        rows = list(csv.DictReader(io.StringIO(text)))
        return self._make_preview(user_id, rows, filename)

    def preview_xlsx(self, user_id: str, path: Path) -> ImportPreview:
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:
            raise ValueError("缺少 openpyxl，无法读取 xlsx") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            raise ValueError("表格为空")
        timetable_rows = parse_bupt_timetable_rows(values)
        if timetable_rows:
            return self._make_preview(user_id, timetable_rows, path.name)
        headers = [str(item or "").strip() for item in values[0]]
        rows = [dict(zip(headers, row, strict=False)) for row in values[1:] if any(row)]
        return self._make_preview(user_id, rows, path.name)

    def preview_xls(self, user_id: str, path: Path) -> ImportPreview:
        try:
            import xlrd
        except ModuleNotFoundError as exc:
            raise ValueError("缺少 xlrd，无法读取教务系统 xls 课表") from exc
        workbook = xlrd.open_workbook(path)
        sheet = workbook.sheet_by_index(0)
        values = [sheet.row_values(index) for index in range(sheet.nrows)]
        rows = parse_bupt_timetable_rows(values)
        if not rows:
            raise ValueError("没有识别到教务系统课表；请使用教务系统直接下载的个人课表")
        return self._make_preview(user_id, rows, path.name)

    def confirm(self, user_id: str, token: str) -> tuple[str, int]:
        preview = self._previews.pop(token, None)
        if preview is None or preview.user_id != str(user_id):
            raise ValueError("导入 token 无效或不属于当前用户")
        batch_id = secrets.token_hex(8)
        count = 0
        for row in preview.rows:
            self.add(
                user_id,
                str(row["name"]),
                int(row["weekday"]),
                (int(row["start_section"]), int(row["end_section"])),
                str(row["weeks"]),
                str(row.get("location", "")),
                str(row.get("teacher", "")),
                source=preview.source,
                batch_id=batch_id,
            )
            count += 1
        return batch_id, count

    def preview_rows(self, user_id: str, rows: list[dict[str, Any]], source: str) -> ImportPreview:
        return self._make_preview(user_id, rows, source)

    def _make_preview(self, user_id: str, rows: list[dict[str, Any]], source: str) -> ImportPreview:
        if not rows:
            raise ValueError("没有可导入的数据行")
        mapping = self._column_mapping(rows[0].keys())
        required = {"name", "weekday", "start_section", "end_section", "weeks"}
        missing = required - mapping.keys()
        if missing:
            raise ValueError("无法确定列映射：" + "、".join(sorted(missing)))
        normalized = []
        for raw in rows:
            item = {target: raw.get(source_name, "") for target, source_name in mapping.items()}
            item["weekday"] = parse_weekday(str(item["weekday"]))
            item["start_section"] = int(item["start_section"])
            item["end_section"] = int(item["end_section"])
            item["weeks"] = str(item["weeks"])
            normalized.append(item)
        token = secrets.token_urlsafe(8)
        preview = ImportPreview(token, str(user_id), normalized, source)
        self._previews[token] = preview
        return preview

    def _column_mapping(self, headers) -> dict[str, str]:
        result: dict[str, str] = {}
        for header in headers:
            normalized = str(header).strip().lower()
            for target, aliases in self.COLUMN_ALIASES.items():
                if normalized in {item.lower() for item in aliases}:
                    result[target] = str(header)
        return result

    @staticmethod
    def _from_row(row) -> CourseRecord:
        return CourseRecord(
            int(row["course_id"]),
            str(row["name"]),
            int(row["weekday"]),
            int(row["start_section"]),
            int(row["end_section"]),
            str(row["weeks"]),
            str(row["location"]),
            str(row["teacher"]),
            int(row["reminder_minutes"]) if row["reminder_minutes"] is not None else None,
        )


def parse_weekday(value: str) -> int:
    text = value.strip().replace("星期", "").replace("周", "")
    values = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5, "六": 6, "日": 7, "天": 7}
    if text in values:
        return values[text]
    number = int(text)
    if not 1 <= number <= 7:
        raise ValueError("星期必须为 1～7")
    return number


def parse_sections(value: str) -> tuple[int, int]:
    match = re.fullmatch(r"(\d+)(?:[-~～](\d+))?", value.strip())
    if not match:
        raise ValueError("节次格式示例：1-2")
    return int(match.group(1)), int(match.group(2) or match.group(1))


def normalize_weeks(value: str) -> set[int]:
    result: set[int] = set()
    text = value.strip().replace("周", "")
    for part in re.split(r"[,，]", text):
        match = re.fullmatch(r"(\d+)(?:[-~～](\d+))?", part.strip())
        if not match:
            raise ValueError("周次格式示例：1-16 或 1-8,10-16")
        start, end = int(match.group(1)), int(match.group(2) or match.group(1))
        if not 1 <= start <= end <= 30:
            raise ValueError("周次必须在 1～30")
        result.update(range(start, end + 1))
    return result


def parse_bupt_timetable_rows(values: list | tuple) -> list[dict[str, Any]]:
    """Parse the matrix timetable exported by the BUPT academic system."""

    header_index = -1
    weekday_columns: dict[int, int] = {}
    for row_index, row in enumerate(values):
        candidates: dict[int, int] = {}
        for column, value in enumerate(row):
            text = str(value or "").strip()
            if re.fullmatch(r"(?:星期|周)[一二三四五六日天]", text):
                candidates[column] = parse_weekday(text)
        if len(candidates) >= 5:
            header_index = row_index
            weekday_columns = candidates
            break
    if header_index < 0:
        return []

    pattern = re.compile(
        r"(?P<name>[^\n]+)\n(?P<teacher>[^\n]+)\n"
        r"(?P<weeks>\d+(?:[-~～,，]\d+)*)\[周\]\n"
        r"(?P<location>[^\n]*)\n\[(?P<sections>\d+(?:-\d+)*)\]节"
    )
    rows: list[dict[str, Any]] = []
    seen: set[tuple] = set()
    for row in values[header_index + 1 :]:
        for column, weekday in weekday_columns.items():
            if column >= len(row):
                continue
            text = str(row[column] or "").replace("\r\n", "\n").replace("\r", "\n").strip()
            for match in pattern.finditer(text):
                sections = [int(value) for value in match.group("sections").split("-")]
                item = {
                    "name": match.group("name").strip(),
                    "teacher": match.group("teacher").strip(),
                    "location": match.group("location").strip(),
                    "weekday": weekday,
                    "start_section": min(sections),
                    "end_section": max(sections),
                    "weeks": match.group("weeks").replace("～", "-").replace("~", "-"),
                }
                key = tuple(item.values())
                if key not in seen:
                    seen.add(key)
                    rows.append(item)
    return rows


def _section_times() -> dict[int, time]:
    defaults = {
        1: "08:00",
        2: "08:50",
        3: "09:50",
        4: "10:40",
        5: "11:30",
        6: "13:30",
        7: "14:20",
        8: "15:20",
        9: "16:10",
        10: "17:00",
        11: "18:30",
        12: "19:20",
        13: "20:10",
    }
    raw = os.getenv("AMADEUS_SECTION_TIMES", "").strip()
    if raw:
        try:
            defaults.update({int(key): str(value) for key, value in json.loads(raw).items()})
        except (ValueError, TypeError, json.JSONDecodeError):
            pass
    return {key: time.fromisoformat(value) for key, value in defaults.items()}
